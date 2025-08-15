import pandas as pd
from openpyxl import load_workbook

def convert_to_date(value):
    try:
        return pd.to_datetime(value, errors="coerce", dayfirst=True)
    except:
        return value


def add_resolution_period_column_logic(input_excel):
    # Load the Excel file
    original_wb = load_workbook(input_excel)
    original_sheet_name = original_wb.sheetnames[0]

    # Read the sheet into a DataFrame
    df = pd.read_excel(input_excel, engine="openpyxl", sheet_name=original_sheet_name)

    # Convert column names to lowercase
    df.columns = [col.lower() for col in df.columns]

    # Convert 'created' and 'resolved' columns to proper date formats
    if "created" in df.columns:
        df["created"] = df["created"].apply(convert_to_date)
    if "resolved" in df.columns:
        df["resolved"] = df["resolved"].apply(convert_to_date)

    # Add the 'days spent to resolve' column
    if "created" in df.columns and "resolved" in df.columns:
        df["days spent to resolve"] = (df["resolved"] - df["created"]).dt.days
    else:
        raise ValueError("Required columns 'Created' and 'Resolved' are missing.")

    # Save the updated DataFrame to a new Excel file
    output_excel = input_excel.replace(".xlsx", "_with_resolution_period.xlsx")
    with pd.ExcelWriter(output_excel, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=original_sheet_name)

    return output_excel
    