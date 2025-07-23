import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment, Font
import re

DATE_COLUMNS = ["Created", "Updated", "Last Viewed", "Resolved"]
SPECIAL_DATE_COLUMNS = ["Attachment", "Comment"]

def convert_to_date(value):
    try:
        return pd.to_datetime(value, errors="coerce", dayfirst=True)
    except:
        return value

def convert_special_date_format(value):
    if pd.isna(value) or str(value).strip().lower() == "nan":
        return ""  # Return an empty string for blank cells
    parts = str(value).split()
    try:
        date_part = parts[0]
        formatted_date = pd.to_datetime(date_part, format="%d.%m.%Y", errors="coerce", dayfirst=True)
        if pd.isna(formatted_date):
            return value  # Return original if date parsing fails
        return formatted_date.strftime("%Y-%m-%d") + (" " + " ".join(parts[1:]) if len(parts) > 1 else "")
    except:
        return value

# Open file dialog to choose an Excel file
root = tk.Tk()
root.withdraw()
file_path = filedialog.askopenfilename(title="Select a Defect Report", filetypes=[("Excel Files", "*.xlsx;*.xls")])


if file_path:
    # Load original sheet name
    original_wb = load_workbook(file_path)
    original_sheet_name = original_wb.sheetnames[0]
    table_name = re.sub(r'\W+', '', original_sheet_name)  # Sanitize table name

    # Read Excel sheet into DataFrame
    df = pd.read_excel(file_path, engine="openpyxl", sheet_name=original_sheet_name)

    # Convert date columns
    for col in DATE_COLUMNS:
        if col in df.columns:
            df[col] = df[col].apply(convert_to_date)

    for col in df.columns:
        if any(keyword in col for keyword in SPECIAL_DATE_COLUMNS):
            df[col] = df[col].astype(str).apply(convert_special_date_format)

    # Save converted DataFrame to new Excel file
    new_file_path = file_path.replace(".xlsx", "_updated.xlsx")
    with pd.ExcelWriter(new_file_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=original_sheet_name)

    # Re-open the file for formatting
    wb = load_workbook(new_file_path)
    ws = wb[original_sheet_name]

    # Create Excel table with green style
    table_ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    table = Table(displayName=table_name, ref=table_ref)

    style = TableStyleInfo(
        name="TableStyleMedium11",  # Green style
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # Align all cells and apply wrap text to non-header cells
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if cell.row == 1: # Header only
                cell.alignment = Alignment(vertical="center")
                cell.font = Font(color="FFFFFF", bold=True)
            else:
                cell.alignment = Alignment(vertical="center", wrap_text=True)

    # Auto-fit column widths
    for col_num in range(1, ws.max_column + 1):
        max_length = 0
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_num, max_col=col_num):
            for cell in row:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
        adjusted_width = min(max_length + 2, 50)
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = adjusted_width

    # Save formatted workbook
    wb.save(new_file_path)
    messagebox.showinfo("Success", f"File saved with proper data format:\n{new_file_path}")

else:
    messagebox.showwarning("No Selection", "No file selected.")
